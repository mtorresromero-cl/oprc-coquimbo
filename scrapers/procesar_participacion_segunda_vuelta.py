"""Participación electoral en las 6 segundas vueltas (presidencial y
gobernador regional) que faltaban en participacion_electoral: esa tabla se
construyó desde extras/participación/ (csv ya agregados por comuna que el
usuario descargó de SERVEL), y esos csv nunca incluyeron una segunda
vuelta — solo bundlean el día de la primera vuelta con las demás
elecciones simultáneas de esa misma jornada.

Los xlsx de resultados en extras/elecciones/ (usados por
procesar_elecciones_recientes.py para candidato-por-candidato) traen,
además de la hoja de resultados, una hoja de participación a nivel de
MESA para la misma elección — nunca explotada hasta ahora. Se agrega
sumando inscritos/votos por comuna, igual que procesar_elecciones_recientes.py
hace para los plebiscitos (también mesa-level).

El nombre de la hoja varía: "Participación" en los archivos que solo
reportan votación en Chile (gobernador, que no vota en el extranjero),
"Participación en Chile" en los presidenciales (que sí traen extranjero
aparte, en su propia hoja "Participación en el extranjero" — no se usa
acá: el padrón/participación de Coquimbo es solo voto dentro de Chile).
La columna con el conteo de votos también varía: "Votación" en los
presidenciales, "Votos" en los gobernadores.
"""

import json
import sqlite3
from pathlib import Path

import openpyxl
from base import BaseScraper
from procesar_elecciones_recientes import COMUNAS_NOMBRE

ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
RAW_DIR = ROOT / "extras" / "elecciones"

# (archivo, jornada, año, etiqueta, tipos_relacionados en resultado_electoral)
JORNADAS = [
    (
        "2013_presidencial_2V.xlsx",
        "presidencial_2v_2013",
        2013,
        "Presidencial (2ª vuelta) 2013",
        ["presidencial_2v"],
    ),
    (
        "2017_presidencial_2V.xlsx",
        "presidencial_2v_2017",
        2017,
        "Presidencial (2ª vuelta) 2017",
        ["presidencial_2v"],
    ),
    (
        "2021_06_gobernadores_2V.xlsx",
        "gobernador_2v_2021",
        2021,
        "Gobernador Regional (2ª vuelta) 2021",
        ["gobernador_2v"],
    ),
    (
        "2021_12_presidencial_2V.xlsx",
        "presidencial_2v_2021",
        2021,
        "Presidencial (2ª vuelta) 2021",
        ["presidencial_2v"],
    ),
    (
        "2024_11_gobernadores_2V.xlsx",
        "gobernador_2v_2024",
        2024,
        "Gobernador Regional (2ª vuelta) 2024",
        ["gobernador_2v"],
    ),
    (
        "2025_presidencial_2V.xlsx",
        "presidencial_2v_2025",
        2025,
        "Presidencial (2ª vuelta) 2025",
        ["presidencial_2v"],
    ),
]


class ProcesadorParticipacionSegundaVuelta(BaseScraper):
    """No es un scraper web: lee la hoja de participación a nivel de mesa de
    los xlsx de segunda vuelta que el usuario ya descargó de SERVEL
    (extras/elecciones/) y la agrega por comuna. Se ejecuta una sola vez
    por archivo nuevo."""

    nombre = "procesar_participacion_segunda_vuelta"
    frecuencia = "una_vez"

    def recolectar(self) -> list[dict]:
        registros = []
        for archivo, jornada, anno, etiqueta, tipos in JORNADAS:
            ruta = RAW_DIR / archivo
            if not ruta.exists():
                print(f"[{archivo}] no encontrado, se omite")
                continue
            try:
                filas = self._procesar_archivo(ruta, jornada, anno, etiqueta, tipos)
            except Exception as e:
                print(f"[{archivo}] ERROR: {e}")
                self.stats["errores"] += 1
                continue
            print(f"[{archivo}] {len(filas)} comunas de Coquimbo")
            registros.extend(filas)
        return registros

    def _procesar_archivo(
        self, ruta: Path, jornada: str, anno: int, etiqueta: str, tipos: list[str]
    ) -> list[dict]:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hay_extranjero = "Participación en Chile" in wb.sheetnames
        nombre_hoja = "Participación en Chile" if hay_extranjero else "Participación"
        ws = wb[nombre_hoja]

        encabezados = None
        idx_region = idx_comuna = idx_inscritos = idx_votos = None
        acumulado: dict[str, list[int]] = {}  # comuna_id -> [inscritos, votos]

        for row in ws.iter_rows(values_only=True):
            if encabezados is None:
                if row and any(isinstance(c, str) and c.strip() == "Comuna" for c in row):
                    encabezados = [str(c).strip() if c else "" for c in row]
                    idx_region = next(
                        (i for i, h in enumerate(encabezados) if h.lower() in ("región", "region")),
                        None,
                    )
                    idx_comuna = encabezados.index("Comuna")
                    idx_inscritos = next(
                        (i for i, h in enumerate(encabezados) if h == "Inscritos"), None
                    )
                    votos_labels = ("votación", "votos")
                    idx_votos = next(
                        (i for i, h in enumerate(encabezados) if h.lower() in votos_labels), None
                    )
                continue

            if idx_region is not None:
                region_val = row[idx_region]
                if not region_val or "COQUIMBO" not in str(region_val).strip().upper():
                    continue

            comuna_txt = str(row[idx_comuna] or "").strip().upper()
            comuna_id = COMUNAS_NOMBRE.get(comuna_txt)
            if comuna_id is None:
                continue

            try:
                inscritos = int(row[idx_inscritos])
                votos = int(row[idx_votos])
            except (TypeError, ValueError):
                continue

            actual = acumulado.setdefault(comuna_id, [0, 0])
            actual[0] += inscritos
            actual[1] += votos

        wb.close()

        registros = []
        for comuna_id, (inscritos, votantes) in acumulado.items():
            pct = round(votantes / inscritos * 100, 2) if inscritos else 0.0
            registros.append(
                {
                    "jornada": jornada,
                    "anno": anno,
                    "etiqueta": etiqueta,
                    "tipos_relacionados": ",".join(tipos),
                    "comuna_id": comuna_id,
                    "inscritos": inscritos,
                    "votantes": votantes,
                    "participacion_pct": pct,
                }
            )
        return registros

    def procesar(self, registros: list[dict]) -> list[dict]:
        return registros

    def guardar(self, registros: list[dict]) -> None:
        jornadas = {r["jornada"] for r in registros}
        for jornada in jornadas:
            self.db.execute("DELETE FROM participacion_electoral WHERE jornada = ?", (jornada,))
        for r in registros:
            self.db.execute(
                """
                INSERT INTO participacion_electoral
                    (jornada, anno, etiqueta, tipos_relacionados, comuna_id,
                     inscritos, votantes, participacion_pct)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    r["jornada"], r["anno"], r["etiqueta"], r["tipos_relacionados"],
                    r["comuna_id"], r["inscritos"], r["votantes"], r["participacion_pct"],
                ),
            )
            self.stats["nuevos"] += 1
        self.db.commit()

    def exportar_json(self) -> None:
        self.db.row_factory = sqlite3.Row
        filas = self.db.execute(
            """
            SELECT jornada, anno, etiqueta, tipos_relacionados, comuna_id,
                   inscritos, votantes, participacion_pct
            FROM participacion_electoral
            ORDER BY anno DESC, jornada, comuna_id
            """
        ).fetchall()

        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        (PROCESSED_DIR / "participacion-electoral.json").write_text(
            json.dumps([dict(f) for f in filas], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Exportados {len(filas)} registros de participación a {PROCESSED_DIR}")


if __name__ == "__main__":
    proc = ProcesadorParticipacionSegundaVuelta()
    proc.ejecutar()
    print("Estadísticas:", proc.stats)
