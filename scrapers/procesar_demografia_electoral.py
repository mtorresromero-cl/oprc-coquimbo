"""Composición demográfica (sexo, rango etario, nacionalidad) de quienes
efectivamente votaron, por jornada y comuna — a partir de la hoja
"Descripción votantes" que trae cada xlsx de extras/elecciones/ junto a los
resultados (ver scrapers/procesar_elecciones_recientes.py), nivel mesa.

Esta hoja describe a quién votó ESE DÍA, no por qué papeleta — es la misma
para todas las elecciones simultáneas de una misma jornada (ej. alcaldes y
concejales en mayo 2021 comparten el mismo electorado que fue a votar), así
que se extrae de un solo archivo representante por jornada en vez de los 39.

plebiscito_2020 queda sin datos: ninguno de sus dos archivos trae la hoja
para votantes en Chile, solo para el voto en el extranjero — hueco real de
la fuente, no un error de procesamiento (confirmado revisando
wb.sheetnames de ambos archivos).
"""

import json
import sqlite3
from pathlib import Path

import openpyxl
from base import BaseScraper
from procesar_elecciones_recientes import COMUNAS_NOMBRE

ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
RAW_DIR = ROOT / "extras" / "elecciones"

# (jornada, año, archivo representante de esa jornada)
JORNADA_ARCHIVO = [
    ("municipal_2012", 2012, "2012_alcaldes.xlsx"),
    ("parlamentaria_presidencial_2013", 2013, "2013_diputados.xlsx"),
    ("municipal_2016", 2016, "2016_alcaldes.xlsx"),
    ("parlamentaria_presidencial_2017", 2017, "2017_diputados.xlsx"),
    ("megaeleccion_2021", 2021, "2021_05_alcaldes.xlsx"),
    ("parlamentaria_presidencial_2021", 2021, "2021_11_diputados.xlsx"),
    ("plebiscito_2022", 2022, "2022_PlebiscitoConstitucional.xlsx"),
    ("plebiscito_2023", 2023, "2023_PlebiscitoConstitucional.xlsx"),
    ("municipal_2024", 2024, "2024_alcaldes.xlsx"),
    ("parlamentaria_presidencial_2025", 2025, "2025_diputados.xlsx"),
]


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


class ProcesadorDemografiaElectoral(BaseScraper):
    """No es un scraper web: lee la hoja "Descripción votantes" de los xlsx
    que el usuario ya descargó de SERVEL (extras/elecciones/, mismos
    archivos que procesar_elecciones_recientes.py)."""

    nombre = "procesar_demografia_electoral"
    frecuencia = "una_vez"

    def recolectar(self) -> list[dict]:
        registros = []
        for jornada, anno, archivo in JORNADA_ARCHIVO:
            ruta = RAW_DIR / archivo
            if not ruta.exists():
                print(f"[{archivo}] no encontrado, se omite")
                continue
            try:
                filas = self._procesar_archivo(ruta, jornada, anno)
            except Exception as e:
                print(f"[{archivo}] ERROR: {e}")
                self.stats["errores"] += 1
                continue
            print(f"[{jornada}] {len(filas)} combinaciones comuna/sexo/rango/nacionalidad")
            registros.extend(filas)
        return registros

    def _procesar_archivo(self, ruta: Path, jornada: str, anno: int) -> list[dict]:
        wb = openpyxl.load_workbook(ruta, read_only=True)
        nombre_hoja = next(
            (
                h
                for h in wb.sheetnames
                if "descripci" in h.lower()
                and "votant" in h.lower()
                and "extranjero" not in h.lower()
            ),
            None,
        )
        if nombre_hoja is None:
            return []
        ws = wb[nombre_hoja]

        encabezado = None
        idx_encabezado = None
        for i, fila in enumerate(ws.iter_rows(max_row=15, values_only=True)):
            textos = [_texto(c).lower() for c in fila]
            if "región" in textos:
                encabezado = textos
                idx_encabezado = i
                break
        if encabezado is None:
            return []

        col = {nombre: i for i, nombre in enumerate(encabezado) if nombre}
        idx_region = col.get("región")
        idx_comuna = col.get("comuna")
        idx_sexo = col.get("sexo")
        idx_rango = col.get("rango etario")
        idx_nacionalidad = col.get("nacionalidad")
        idx_votantes = col.get("votantes") or col.get("votación")
        if None in (idx_region, idx_comuna, idx_sexo, idx_rango, idx_nacionalidad, idx_votantes):
            return []

        acumulado: dict[tuple, int] = {}
        for fila in ws.iter_rows(min_row=idx_encabezado + 2, values_only=True):
            if not fila or fila[idx_region] is None:
                continue
            region_txt = _texto(fila[idx_region]).upper()
            if "COQUIMBO" not in region_txt:
                continue
            comuna_txt = _texto(fila[idx_comuna]).upper()
            comuna_id = COMUNAS_NOMBRE.get(comuna_txt)
            if comuna_id is None:
                continue
            votantes = fila[idx_votantes]
            if votantes is None or not isinstance(votantes, (int, float)):
                continue

            sexo = _texto(fila[idx_sexo]).upper()
            rango = _texto(fila[idx_rango])
            extranjero = _texto(fila[idx_nacionalidad]).upper() != "CHILE"
            clave = (comuna_id, sexo, rango, extranjero)
            acumulado[clave] = acumulado.get(clave, 0) + int(votantes)

        return [
            {
                "jornada": jornada,
                "anno": anno,
                "comuna_id": comuna_id,
                "sexo": sexo,
                "rango_etario": rango,
                "extranjero": extranjero,
                "votantes": votantes,
            }
            for (comuna_id, sexo, rango, extranjero), votantes in acumulado.items()
        ]

    def procesar(self, registros: list[dict]) -> list[dict]:
        return registros

    def guardar(self, registros: list[dict]) -> None:
        jornadas = {r["jornada"] for r in registros}
        for jornada in jornadas:
            self.db.execute("DELETE FROM padron_demografico WHERE jornada = ?", (jornada,))
        for r in registros:
            self.db.execute(
                """
                INSERT INTO padron_demografico
                    (jornada, anno, comuna_id, sexo, rango_etario, extranjero, votantes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    r["jornada"], r["anno"], r["comuna_id"], r["sexo"],
                    r["rango_etario"], r["extranjero"], r["votantes"],
                ),
            )
            self.stats["nuevos"] += 1
        self.db.commit()

    def exportar_json(self) -> None:
        self.db.row_factory = sqlite3.Row
        filas = self.db.execute(
            """
            SELECT jornada, anno, comuna_id, sexo, rango_etario, extranjero, votantes
            FROM padron_demografico
            ORDER BY anno, comuna_id, sexo, rango_etario
            """
        ).fetchall()

        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        (PROCESSED_DIR / "padron-demografico.json").write_text(
            json.dumps([dict(f) for f in filas], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Exportados {len(filas)} registros demográficos a {PROCESSED_DIR}")


if __name__ == "__main__":
    proc = ProcesadorDemografiaElectoral()
    proc.ejecutar()
    print("Estadísticas:", proc.stats)
