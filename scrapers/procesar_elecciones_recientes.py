"""Resultados electorales 2012-2025 de la Región de Coquimbo, a partir de
los archivos xlsx nacionales que descargó el usuario directamente de
SERVEL (extras/elecciones/, no versionado en git por tamaño — ver
.gitignore). A diferencia de 1989-2009 (scrapers/servel_historico.py, sitio
legado sin archivos descargables) o de las elecciones 2012+ mostradas en el
sitio actual como dashboards de Power BI (no exportables), estos xlsx sí
traen una hoja "Votación por comuna" con el detalle candidato por
candidato, nacional — este script filtra a Región de Coquimbo y normaliza
al mismo esquema que usa resultado_electoral (compartida con
servel_historico.py).

Cada archivo tiene columnas ligeramente distintas según el tipo de
elección (presidencial no trae Partido; CCPI trae "Pueblo" en vez de
Lista/Pacto/Partido; gobernadores 2021 tiene un typo real "Carrgo" en vez
de "Cargo") — las columnas se detectan por nombre de encabezado, no por
índice fijo. Los plebiscitos son la excepción estructural: la hoja trae
el detalle por MESA, no por comuna, así que se agrega sumando votos por
(comuna, opción) antes de guardar.
"""

import json
import sqlite3
from pathlib import Path

import openpyxl
from base import BaseScraper

ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
RAW_DIR = ROOT / "extras" / "elecciones"

# nombre de comuna tal como aparece en los xlsx (MAYÚSCULAS, con o sin
# tilde según el archivo) -> nuestro comuna_id.
COMUNAS_NOMBRE = {
    "LA SERENA": "la-serena",
    "LA HIGUERA": "la-higuera",
    "COQUIMBO": "coquimbo",
    "ANDACOLLO": "andacollo",
    "VICUÑA": "vicuna",
    "VICUNA": "vicuna",
    "PAIHUANO": "paihuano",
    "OVALLE": "ovalle",
    "RIO HURTADO": "rio-hurtado",
    "RÍO HURTADO": "rio-hurtado",
    "MONTE PATRIA": "monte-patria",
    "COMBARBALA": "combarbala",
    "COMBARBALÁ": "combarbala",
    "PUNITAQUI": "punitaqui",
    "ILLAPEL": "illapel",
    "SALAMANCA": "salamanca",
    "LOS VILOS": "los-vilos",
    "CANELA": "canela",
}

# (archivo, año, eleccion_tipo, cargo por defecto)
ARCHIVOS = [
    ("2012_alcaldes.xlsx", 2012, "municipal", "Alcalde"),
    ("2012_concejales.xlsx", 2012, "municipal", "Concejal"),
    ("2013_consejerosregionales.xlsx", 2013, "consejeros_regionales", "Consejero Regional"),
    ("2013_diputados.xlsx", 2013, "diputados", "Diputado"),
    ("2013_presidencial_1V.xlsx", 2013, "presidencial_1v", "Presidente"),
    ("2013_presidencial_2V.xlsx", 2013, "presidencial_2v", "Presidente"),
    ("2013_senatorial.xlsx", 2013, "senadores", "Senador"),
    ("2016_alcaldes.xlsx", 2016, "municipal", "Alcalde"),
    ("2016_concejales.xlsx", 2016, "municipal", "Concejal"),
    ("2017_consejerosregionales.xlsx", 2017, "consejeros_regionales", "Consejero Regional"),
    ("2017_diputados.xlsx", 2017, "diputados", "Diputado"),
    ("2017_presidencial_1V.xlsx", 2017, "presidencial_1v", "Presidente"),
    ("2017_presidencial_2V.xlsx", 2017, "presidencial_2v", "Presidente"),
    ("2017_senatorial.xlsx", 2017, "senadores", "Senador"),
    ("2020_plebiscitoCP.xlsx", 2020, "plebiscito_constitucion", "Opción"),
    ("2020_plebiscitoTipoOrgano.xlsx", 2020, "plebiscito_tipo_organo", "Opción"),
    ("2021_05_CCG.xlsx", 2021, "convencional_constituyente", "Convencional Constituyente"),
    (
        "2021_05_CCPI.xlsx",
        2021,
        "convencional_constituyente_indigena",
        "Convencional Constituyente",
    ),
    ("2021_05_alcaldes.xlsx", 2021, "municipal", "Alcalde"),
    ("2021_05_concejales.xlsx", 2021, "municipal", "Concejal"),
    ("2021_05_gobernadores_1V.xlsx", 2021, "gobernador_1v", "Gobernador Regional"),
    ("2021_06_gobernadores_2V.xlsx", 2021, "gobernador_2v", "Gobernador Regional"),
    ("2021_11_consejerosregionales.xlsx", 2021, "consejeros_regionales", "Consejero Regional"),
    ("2021_11_diputados.xlsx", 2021, "diputados", "Diputado"),
    ("2021_11_presidencial_1V.xlsx", 2021, "presidencial_1v", "Presidente"),
    ("2021_11_senatorial.xlsx", 2021, "senadores", "Senador"),
    ("2021_12_presidencial_2V.xlsx", 2021, "presidencial_2v", "Presidente"),
    ("2022_PlebiscitoConstitucional.xlsx", 2022, "plebiscito_constitucional", "Opción"),
    ("2023_05_CCG.xlsx", 2023, "consejo_constitucional", "Consejero Constitucional"),
    ("2023_05_CCPI.xlsx", 2023, "consejo_constitucional_indigena", "Consejero Constitucional"),
    ("2023_PlebiscitoConstitucional.xlsx", 2023, "plebiscito_constitucional", "Opción"),
    ("2024_10_gobernadores_1V.xlsx", 2024, "gobernador_1v", "Gobernador Regional"),
    ("2024_11_gobernadores_2V.xlsx", 2024, "gobernador_2v", "Gobernador Regional"),
    ("2024_alcaldes.xlsx", 2024, "municipal", "Alcalde"),
    ("2024_concejales.xlsx", 2024, "municipal", "Concejal"),
    ("2024_consejerosregionales.xlsx", 2024, "consejeros_regionales", "Consejero Regional"),
    ("2025_diputados.xlsx", 2025, "diputados", "Diputado"),
    ("2025_presidencial_1V.xlsx", 2025, "presidencial_1v", "Presidente"),
    ("2025_presidencial_2V.xlsx", 2025, "presidencial_2v", "Presidente"),
    ("2025_senadores.xlsx", 2025, "senadores", "Senador"),
]

# nombres de columna que indican quién ganó/quedó nominado — varía según
# el tipo de elección y el archivo (typo real "Carrgo" en gobernadores
# 2021 1V).
COLUMNAS_ELECTO = ["Cargo", "Carrgo", "Nominado", "Selección"]

CANDIDATOS_PLACEHOLDER = {"VOTOS EN BLANCO", "VOTOS NULOS"}


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


class ProcesadorEleccionesRecientes(BaseScraper):
    """No es un scraper web: lee los xlsx que el usuario ya descargó de
    SERVEL (extras/elecciones/) y los normaliza al mismo esquema que
    servel_historico.py. Se ejecuta una sola vez por archivo nuevo, no
    tiene sentido una frecuencia periódica."""

    nombre = "procesar_elecciones_recientes"
    frecuencia = "una_vez"

    def recolectar(self) -> list[dict]:
        registros = []
        for archivo, anno, eleccion_tipo, cargo_defecto in ARCHIVOS:
            ruta = RAW_DIR / archivo
            if not ruta.exists():
                print(f"[{archivo}] no encontrado, se omite")
                continue
            try:
                filas = self._procesar_archivo(ruta, anno, eleccion_tipo, cargo_defecto)
            except Exception as e:
                print(f"[{archivo}] ERROR: {e}")
                self.stats["errores"] += 1
                continue
            print(f"[{archivo}] {len(filas)} registros de Coquimbo")
            registros.extend(filas)
        return registros

    def _procesar_archivo(
        self, ruta: Path, anno: int, eleccion_tipo: str, cargo_defecto: str
    ) -> list[dict]:
        wb = openpyxl.load_workbook(ruta, read_only=True)
        # el nombre de la hoja varía entre archivos ("Votación por comuna",
        # "Votación por comuna en Chile", "Votación comunal en Chile" en
        # los plebiscitos 2022) — se detecta por contener "comuna" +
        # "votaci", no por coincidencia exacta.
        nombre_hoja = next(
            (
                h
                for h in wb.sheetnames
                if "comuna" in h.lower() and "votaci" in h.lower() and "extranjero" not in h.lower()
            ),
            None,
        )
        if nombre_hoja is None:
            return []
        ws = wb[nombre_hoja]

        encabezado = None
        idx_encabezado = None
        for i, fila in enumerate(ws.iter_rows(max_row=15, values_only=True)):
            textos = [_texto(c) for c in fila]
            if "Región" in textos or "Nro.Región" in textos:
                encabezado = textos
                idx_encabezado = i
                break
        if encabezado is None:
            return []

        col = {nombre: i for i, nombre in enumerate(encabezado) if nombre}
        idx_region = col.get("Región")
        idx_comuna = col.get("Comuna")
        idx_votos = col.get("Votos")
        if idx_region is None or idx_comuna is None or idx_votos is None:
            return []

        idx_nombres = col.get("Nombres")
        idx_ap1 = col.get("Primer apellido")
        idx_ap2 = col.get("Segundo apellido")
        idx_candidato = col.get("Candidato")  # 2023 CCG/CCPI: nombre completo en 1 columna
        idx_opcion = col.get("Opción") or col.get("Opciones")
        idx_partido = col.get("Partido")
        idx_pacto = col.get("Pacto")
        idx_pueblo = col.get("Pueblo")
        idx_mesa = col.get("Mesa")
        idx_electo = next((col[c] for c in COLUMNAS_ELECTO if c in col), None)

        es_mesa = idx_mesa is not None  # plebiscitos: nivel mesa, hay que agregar

        def celda(fila, idx) -> str:
            if idx is None or idx >= len(fila):
                return ""
            return _texto(fila[idx])

        acumulado: dict[tuple, dict] = {}  # solo se usa si es_mesa
        registros: list[dict] = []

        for fila in ws.iter_rows(min_row=idx_encabezado + 2, values_only=True):
            if not fila or fila[idx_region] is None:
                continue
            region_txt = _texto(fila[idx_region]).upper()
            if "COQUIMBO" not in region_txt:
                continue
            comuna_txt = _texto(fila[idx_comuna]).upper()
            comuna_id = COMUNAS_NOMBRE.get(comuna_txt)
            if comuna_id is None:
                continue

            if idx_nombres is not None:
                partes = (celda(fila, idx_nombres), celda(fila, idx_ap1), celda(fila, idx_ap2))
                nombre_completo = " ".join(p for p in partes if p).strip()
            elif idx_candidato is not None:
                nombre_completo = celda(fila, idx_candidato)
            elif idx_opcion is not None:
                nombre_completo = celda(fila, idx_opcion)
            else:
                continue

            if not nombre_completo or nombre_completo.upper() in CANDIDATOS_PLACEHOLDER:
                continue

            votos = fila[idx_votos]
            if votos is None or not isinstance(votos, (int, float)):
                continue
            votos = int(votos)

            texto_electo = celda(fila, idx_electo)
            partido = celda(fila, idx_partido) or celda(fila, idx_pueblo)  # CCPI: pueblo = partido
            pacto = celda(fila, idx_pacto)

            registro = {
                "eleccion_tipo": eleccion_tipo,
                "anno": anno,
                "comuna_id": comuna_id,
                "candidato": nombre_completo,
                "partido": partido or None,
                "pacto": pacto or None,
                "votos": votos,
                "electo": bool(texto_electo),
                "cargo": cargo_defecto,
                "fuente_url": None,  # archivo local, no una URL pública
            }

            if es_mesa:
                clave = (comuna_id, nombre_completo)
                if clave in acumulado:
                    acumulado[clave]["votos"] += votos
                    acumulado[clave]["electo"] = acumulado[clave]["electo"] or bool(texto_electo)
                else:
                    acumulado[clave] = registro
            else:
                registros.append(registro)

        if es_mesa:
            registros = list(acumulado.values())

        # porcentaje: no viene en el archivo, se calcula sobre el total de
        # la comuna dentro de esta misma elección.
        totales_comuna: dict[str, int] = {}
        for r in registros:
            totales_comuna[r["comuna_id"]] = totales_comuna.get(r["comuna_id"], 0) + r["votos"]
        for r in registros:
            total = totales_comuna.get(r["comuna_id"], 0)
            r["porcentaje"] = round(r["votos"] / total * 100, 2) if total else None

        return registros

    def procesar(self, registros: list[dict]) -> list[dict]:
        return registros

    def guardar(self, registros: list[dict]) -> None:
        combinaciones = {(r["eleccion_tipo"], r["anno"]) for r in registros}
        for eleccion_tipo, anno in combinaciones:
            self.db.execute(
                "DELETE FROM resultado_electoral WHERE eleccion_tipo = ? AND anno = ?",
                (eleccion_tipo, anno),
            )
        for r in registros:
            self.db.execute(
                """
                INSERT INTO resultado_electoral
                    (eleccion_tipo, anno, comuna_id, candidato, partido, pacto,
                     votos, porcentaje, electo, cargo, fuente_url)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    r["eleccion_tipo"], r["anno"], r["comuna_id"], r["candidato"],
                    r["partido"], r["pacto"], r["votos"], r["porcentaje"],
                    r["electo"], r["cargo"], r["fuente_url"],
                ),
            )
            self.stats["nuevos"] += 1
        self.db.commit()

    def exportar_json(self) -> None:
        self.db.row_factory = sqlite3.Row
        filas = self.db.execute(
            """
            SELECT eleccion_tipo, anno, comuna_id, candidato, partido, pacto,
                   votos, porcentaje, electo, cargo, fuente_url
            FROM resultado_electoral
            WHERE anno >= 2012
            ORDER BY anno DESC, eleccion_tipo, comuna_id, votos DESC
            """
        ).fetchall()

        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        (PROCESSED_DIR / "resultados-electorales-recientes.json").write_text(
            json.dumps([dict(f) for f in filas], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Exportados {len(filas)} resultados electorales recientes a {PROCESSED_DIR}")


if __name__ == "__main__":
    proc = ProcesadorEleccionesRecientes()
    proc.ejecutar()
    print("Estadísticas:", proc.stats)
